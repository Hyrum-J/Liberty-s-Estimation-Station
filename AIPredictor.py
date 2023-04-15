import pandas as pd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_absolute_percentage_error
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import OneHotEncoder

dataset = pd.read_excel("Liberty Data Prediction.xlsx", 'Important Data')
path = "LibertyPriceBook.xlsx"
wb = openpyxl.load_workbook(path)

permitPrediction = LinearRegression()
LConcretePrediction = LinearRegression()
MConcretePrediction = LinearRegression()
LBrickPrediction = LinearRegression()
MBrickPrediction = LinearRegression()
ElectricalPrediction = LinearRegression()
FramingPrediction = LinearRegression()
LumberPrediction = LinearRegression()
PlumbingPrediction = LinearRegression()
LRoofingPrediction = LinearRegression()
MRoofingPrediction = LinearRegression()
LSheetrockPrediction = LinearRegression()
MSheetrockPrediction = LinearRegression()
LSidingPrediction = LinearRegression()
MSidingPrediction = LinearRegression()
LTrimPrediction = LinearRegression()
MTrimPrediction = LinearRegression()
TrussPrediction = LinearRegression()
HeatingPrediction = LinearRegression()
InsulationPrediction = LinearRegression()
EngineeringPrediction = LinearRegression()
ExcavationPrediction = LinearRegression()
FoundationPrediction = LinearRegression()
GravelPrediction = LinearRegression()
PaintingPrediction = LinearRegression()

def train():
    global permitPrediction
    global LConcretePrediction 
    global MConcretePrediction 
    global LBrickPrediction 
    global MBrickPrediction 
    global ElectricalPrediction
    global FramingPrediction
    global LumberPrediction
    global PlumbingPrediction
    global LRoofingPrediction
    global MRoofingPrediction
    global LSheetrockPrediction
    global MSheetrockPrediction
    global LSidingPrediction
    global MSidingPrediction
    global LTrimPrediction
    global MTrimPrediction
    global TrussPrediction
    global HeatingPrediction
    global InsulationPrediction
    global EngineeringPrediction
    global ExcavationPrediction
    global FoundationPrediction
    global GravelPrediction 
    global PaintingPrediction

    X = dataset.drop(['Permit/SqFt','CLabor/SqFt','CMaterials/SqFt','Blabor/SqFt','Bmaterials/SqFt','Electrical/SqFt','Framing/SqFt','Lumber/SqFt','Plumbing/SqFt','RLabor/SqFt','Rmaterials/SqFt','SLabor/SqFt','SMaterials/SqFt','SiLabor/SqFt','SiMaterials/SqFt','TLabor/SqFt','TMaterials/SqFt','Truss/SqFt','Heating/SqFt','Insulation/SqFt','Engineering/SqFt','Excavation/SqFt','Foundation/SqFt','Gravel/SqFt','Painting/SqFt'], axis=1)
    YP = dataset['Permit/SqFt']
    YCL = dataset['CLabor/SqFt']
    YCM = dataset['CMaterials/SqFt']
    YBL = dataset['Blabor/SqFt']
    YBM = dataset['Bmaterials/SqFt']
    YEL = dataset['Electrical/SqFt']
    YFR = dataset['Framing/SqFt']
    YL = dataset['Lumber/SqFt']
    YPL = dataset['Plumbing/SqFt']
    YRL = dataset['RLabor/SqFt']
    YRM = dataset['Rmaterials/SqFt']
    YSL = dataset['SLabor/SqFt']
    YSM = dataset['SMaterials/SqFt']
    YSIL = dataset['SiLabor/SqFt']
    YSIM = dataset['SiMaterials/SqFt']
    YTL = dataset['TLabor/SqFt']
    YTM = dataset['TMaterials/SqFt']
    YT = dataset['Truss/SqFt']
    YH = dataset['Heating/SqFt']
    YI = dataset['Insulation/SqFt']
    YEN = dataset['Engineering/SqFt']
    YEX = dataset['Excavation/SqFt']
    YFO = dataset['Foundation/SqFt']
    YG = dataset['Gravel/SqFt']
    YPA = dataset['Painting/SqFt']
    

    #Permit
    X_train, X_valid, YP_train, YP_valid = train_test_split(
        X, YP, train_size=0.8, test_size=0.2, random_state=0)
    
    permitPrediction.fit(X_train, YP_train)
    YP_pred = permitPrediction.predict(X_valid)

    #Concrete Labor
    X_train, X_valid, YCL_train, YCL_valid = train_test_split(
        X, YCL, train_size=0.8, test_size=0.2, random_state=0)
    
    LConcretePrediction.fit(X_train, YCL_train)
    YCL_pred = LConcretePrediction.predict(X_valid)

    #Concrete Materials
    X_train, X_valid, YCM_train, YCM_valid = train_test_split(
        X, YCM, train_size=0.8, test_size=0.2, random_state=0)
    
    MConcretePrediction.fit(X_train, YCM_train)
    YCM_pred = MConcretePrediction.predict(X_valid)

    #Brick Labor
    X_train, X_valid, YBL_train, YBL_valid = train_test_split(
        X, YBL, train_size=0.8, test_size=0.2, random_state=0)
    
    LBrickPrediction.fit(X_train, YBL_train)
    YBL_pred = LBrickPrediction.predict(X_valid)

    #Brick Material
    X_train, X_valid, YBM_train, YBM_valid = train_test_split(
        X, YBM, train_size=0.8, test_size=0.2, random_state=0)
    
    MBrickPrediction.fit(X_train, YBM_train)
    YBM_pred = MBrickPrediction.predict(X_valid)

    #Electrical
    X_train, X_valid, YEL_train, YEL_valid = train_test_split(
        X, YEL, train_size=0.8, test_size=0.2, random_state=0)
    
    ElectricalPrediction.fit(X_train, YEL_train)
    YEL_pred = ElectricalPrediction.predict(X_valid)
    
    #Framing
    X_train, X_valid, YFR_train, YFR_valid = train_test_split(
        X, YFR, train_size=0.8, test_size=0.2, random_state=0)
    
    FramingPrediction.fit(X_train, YFR_train)
    YFR_pred = FramingPrediction.predict(X_valid)
    
    #Lumber
    X_train, X_valid, YL_train, YL_valid = train_test_split(
        X, YL, train_size=0.8, test_size=0.2, random_state=0)
    
    LumberPrediction.fit(X_train, YL_train)
    YL_pred = LumberPrediction.predict(X_valid)

    #Plumbing
    X_train, X_valid, YPL_train, YPL_valid = train_test_split(
        X, YPL, train_size=0.8, test_size=0.2, random_state=0)
    
    PlumbingPrediction.fit(X_train, YPL_train)
    YPL_pred = PlumbingPrediction.predict(X_valid)

    #Roofing Labor
    X_train, X_valid, YRL_train, YRL_valid = train_test_split(
        X, YRL, train_size=0.8, test_size=0.2, random_state=0)
    
    LRoofingPrediction.fit(X_train, YRL_train)
    YRL_pred = LRoofingPrediction.predict(X_valid)

    #Roofing Material
    X_train, X_valid, YRM_train, YRM_valid = train_test_split(
        X, YRM, train_size=0.8, test_size=0.2, random_state=0)

    MRoofingPrediction.fit(X_train, YRM_train)
    YRM_pred = MRoofingPrediction.predict(X_valid)

    #Sheetrock Labor    
    X_train, X_valid, YSL_train, YSL_valid = train_test_split(
        X, YSL, train_size=0.8, test_size=0.2, random_state=0)
    
    LSheetrockPrediction.fit(X_train, YSL_train)
    YSL_pred = LSheetrockPrediction.predict(X_valid)
    
    #Sheetrock Material
    X_train, X_valid, YSM_train, YSM_valid = train_test_split(
        X, YSM, train_size=0.8, test_size=0.2, random_state=0)
    
    MSheetrockPrediction.fit(X_train, YSM_train)
    YSM_pred = MSheetrockPrediction.predict(X_valid)

    #Siding Labor
    X_train, X_valid, YSIL_train, YSIL_valid = train_test_split(
        X, YSIL, train_size=0.8, test_size=0.2, random_state=0)
    
    LSidingPrediction.fit(X_train, YSIL_train)
    YSIL_pred = LSidingPrediction.predict(X_valid)

    #Siding Materials
    X_train, X_valid, YSIM_train, YSIM_valid = train_test_split(
        X, YSIM, train_size=0.8, test_size=0.2, random_state=0)
    
    MSidingPrediction.fit(X_train, YSIM_train)
    YSIM_pred = MSidingPrediction.predict(X_valid)

    #Trim Labor
    X_train, X_valid, YTL_train, YTL_valid = train_test_split(
        X, YTL, train_size=0.8, test_size=0.2, random_state=0)
    
    LTrimPrediction.fit(X_train, YTL_train)
    YTL_pred = LTrimPrediction.predict(X_valid)

    #Trim Material
    X_train, X_valid, YTM_train, YTM_valid = train_test_split(
        X, YTM, train_size=0.8, test_size=0.2, random_state=0)
    
    MTrimPrediction.fit(X_train, YTM_train)
    YTM_pred = MTrimPrediction.predict(X_valid)

    #Truss
    X_train, X_valid, YT_train, YT_valid = train_test_split(
        X, YT, train_size=0.8, test_size=0.2, random_state=0)
    
    TrussPrediction.fit(X_train, YT_train)
    YT_pred = TrussPrediction.predict(X_valid)
    
    #Heating
    X_train, X_valid, YH_train, YH_valid = train_test_split(
        X, YH, train_size=0.8, test_size=0.2, random_state=0)
    
    HeatingPrediction.fit(X_train, YH_train)
    YH_pred = HeatingPrediction.predict(X_valid)
    
    #Insulation
    X_train, X_valid, YI_train, YI_valid = train_test_split(
        X, YI, train_size=0.8, test_size=0.2, random_state=0)
    
    InsulationPrediction.fit(X_train, YI_train)
    YI_pred = InsulationPrediction.predict(X_valid)

    #Engineering
    X_train, X_valid, YEN_train, YEN_valid = train_test_split(
        X, YEN, train_size=0.8, test_size=0.2, random_state=0)
    
    EngineeringPrediction.fit(X_train, YEN_train)
    YEN_pred = EngineeringPrediction.predict(X_valid)

    #Excavation
    X_train, X_valid, YEX_train, YEX_valid = train_test_split(
        X, YEX, train_size=0.8, test_size=0.2, random_state=0)
    
    ExcavationPrediction.fit(X_train, YEX_train)
    YEX_pred = ExcavationPrediction.predict(X_valid)

    #Foundation
    X_train, X_valid, YFO_train, YFO_valid = train_test_split(
        X, YFO, train_size=0.8, test_size=0.2, random_state=0)

    FoundationPrediction.fit(X_train, YFO_train)
    YFO_pred = FoundationPrediction.predict(X_valid)

    #Gravel   
    X_train, X_valid, YG_train, YG_valid = train_test_split(
        X, YG, train_size=0.8, test_size=0.2, random_state=0)
    
    GravelPrediction.fit(X_train, YG_train)
    YG_pred = GravelPrediction.predict(X_valid)

    #Painting
    X_train, X_valid, YPA_train, YPA_valid = train_test_split(
        X, YPA, train_size=0.8, test_size=0.2, random_state=0)
    
    PaintingPrediction.fit(X_train, YPA_train)
    YPA_pred = PaintingPrediction.predict(X_valid)
 
    #print(mean_absolute_percentage_error(YP_valid, YP_pred))
    #print(mean_absolute_percentage_error(YCL_valid, YCL_pred))
    #print(mean_absolute_percentage_error(YCM_valid, YCM_pred))
    #print(mean_absolute_percentage_error(YBL_valid, YBL_pred))
    #print(mean_absolute_percentage_error(YBM_valid, YBM_pred))
    #print(mean_absolute_percentage_error(YEL_valid, YEL_pred))
    #print(mean_absolute_percentage_error(YFR_valid, YFR_pred))
    #print(mean_absolute_percentage_error(YL_valid, YL_pred))
    #print(mean_absolute_percentage_error(YPL_valid, YPL_pred))
    #print(mean_absolute_percentage_error(YRL_valid, YRL_pred))
    #print(mean_absolute_percentage_error(YRM_valid, YRM_pred))
    #print(mean_absolute_percentage_error(YTL_valid, YTL_pred))
    #print(mean_absolute_percentage_error(YTM_valid, YTM_pred))
    #print(mean_absolute_percentage_error(YT_valid, YT_pred))
    #print(mean_absolute_percentage_error(YH_valid, YH_pred))
    #print(mean_absolute_percentage_error(YI_valid, YI_pred))
    #print(mean_absolute_percentage_error(YEN_valid, YEN_pred))
    #print(mean_absolute_percentage_error(YEX_valid, YEX_pred))
    #print(mean_absolute_percentage_error(YFO_valid, YFO_pred))
    #print(mean_absolute_percentage_error(YG_valid, YG_pred))
    #print(mean_absolute_percentage_error(YPA_valid, YPA_pred))

def predict(month, day, year, totalSqft, city):
    X_valid = np.array([[month, day, year, totalSqft, city]])
    YP_pred = permitPrediction.predict(X_valid)
    YCL_pred = LConcretePrediction.predict(X_valid)
    YCM_pred = MConcretePrediction.predict(X_valid)
    YBL_pred = LBrickPrediction.predict(X_valid)
    YBM_pred = MBrickPrediction.predict(X_valid)
    YEL_pred = ElectricalPrediction.predict(X_valid)
    YFR_pred = FramingPrediction.predict(X_valid)
    YL_pred = LumberPrediction.predict(X_valid)
    YPL_pred = PlumbingPrediction.predict(X_valid)
    YRL_pred = LRoofingPrediction.predict(X_valid)
    YRM_pred = MRoofingPrediction.predict(X_valid)
    YTL_pred = LTrimPrediction.predict(X_valid)
    YTM_pred = MTrimPrediction.predict(X_valid)
    YT_pred = TrussPrediction.predict(X_valid)
    YH_pred = HeatingPrediction.predict(X_valid)
    YI_pred = InsulationPrediction.predict(X_valid)
    YEN_pred = EngineeringPrediction.predict(X_valid)
    YEX_pred = ExcavationPrediction.predict(X_valid)
    YFO_pred = FoundationPrediction.predict(X_valid)
    YG_pred = GravelPrediction.predict(X_valid)
    YPA_pred = PaintingPrediction.predict(X_valid)
    YSL_pred = LSheetrockPrediction.predict(X_valid)
    YSM_pred = MSheetrockPrediction.predict(X_valid)
    YSIL_pred = LSidingPrediction.predict(X_valid)
    YSIM_pred = MSidingPrediction.predict(X_valid)
        
    ws = wb.active

    ws['B2'] = round(YP_pred[0], 2)
    ws['B3'] = round(YEN_pred[0], 2)
    ws['B4'] = round(YT_pred[0], 2)
    ws['B5'] = round(YRM_pred[0], 2)
    ws['B6'] = round(YRL_pred[0], 2)
    ws['B7'] = round(YBM_pred[0], 2)
    ws['B8'] = round(YBL_pred[0], 2)
    ws['B9'] = round(YL_pred[0], 2)
    ws['B10'] = round(YFR_pred[0], 2)
    ws['B11'] = round(YEL_pred[0], 2)
    ws['B12'] = round(YPL_pred[0], 2)
    ws['B13'] = round(YSM_pred[0], 2)
    ws['B14'] = round(YSL_pred[0], 2)
    ws['B15'] = round(YPA_pred[0], 2)
    ws['B16'] = round(YH_pred[0], 2)
    ws['B17'] = round(YI_pred[0], 2)
    ws['B18'] = round(YTM_pred[0], 2)
    ws['B19'] = round(YTL_pred[0], 2)
    ws['B20'] = round(YCM_pred[0], 2)
    ws['B21'] = round(YCL_pred[0], 2)
    ws['B22'] = round(YFO_pred[0], 2)
    ws['B23'] = round(YEX_pred[0], 2)
    ws['B24'] = round(YG_pred[0], 2)
    ws['B25'] = round(YSIM_pred[0], 2)
    ws['B26'] = round(YSIL_pred[0], 2)
    
    
    wb.save(path)
    

 
